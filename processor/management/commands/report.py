from django.core.management.base import BaseCommand
from django.conf import settings
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
import requests
import os
from collections import defaultdict
from tqdm import tqdm

from processor.models import SCADAData


class Command(BaseCommand):
    help = "Generate SCADA Excel Report (Yearly or Daily)"

    def add_arguments(self, parser):
        parser.add_argument("--year", type=int, help="Generate full year report")
        parser.add_argument("--date", type=str, help="Generate single day report (YYYY-MM-DD)")

    def handle(self, *args, **options):

        year = options.get("year")
        date_str = options.get("date")

        # --------------------------------------------------
        # DATE RANGE
        # --------------------------------------------------
        if year:
            start = timezone.make_aware(datetime(year, 1, 1))
            end = timezone.make_aware(datetime(year + 1, 1, 1))
            report_label = str(year)

        elif date_str:
            try:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                self.stdout.write(self.style.ERROR("Date format must be YYYY-MM-DD"))
                return

            start = timezone.make_aware(date_obj)
            end = start + timedelta(days=1)
            report_label = date_str

        else:
            self.stdout.write(self.style.ERROR("Please provide --year or --date"))
            return

        self.stdout.write(f"\n🚀 Generating SCADA report for {report_label}...\n")

        # --------------------------------------------------
        # FETCH DATA
        # --------------------------------------------------
        queryset = (
            SCADAData.objects
            .filter(datetime__gte=start, datetime__lt=end)
            .order_by("datetime")
        )

        total_records = queryset.count()

        if total_records == 0:
            self.stdout.write(self.style.WARNING("No data found"))
            return

        self.stdout.write(f"📊 Total DB rows fetched: {total_records}")

        # --------------------------------------------------
        # BUILD LOOKUP DICTIONARY
        # --------------------------------------------------
        data_map = defaultdict(dict)

        for record in tqdm(queryset.iterator(),
                           total=total_records,
                           desc="Loading DB Data",
                           unit="rows"):

            loc = str(record.locno).strip()
            data_map[record.datetime][loc] = record

        timestamps = sorted(data_map.keys())
        self.stdout.write(f"🕒 Unique timestamps: {len(timestamps)}")

        # --------------------------------------------------
        # FETCH MACHINE MASTER API
        # --------------------------------------------------
        try:
            machine_api_url = "http://172.16.7.118:8003/api/obs/leaplocs.php"
            response = requests.get(machine_api_url, timeout=10)
            response.raise_for_status()
            machine_list = response.json()
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"API Error: {e}"))
            return

        machine_data = {}
        for item in machine_list:
            api_loc = str(item.get("locno")).strip()
            machine_data[api_loc] = item

        machines = list(machine_data.keys())

        # --------------------------------------------------
        # PARAMETERS (DB FIELD NAME, UNIT, DISPLAY NAME)
        # --------------------------------------------------
        parameters = [
            ("wind_speed", "m/s", "wind_speed"),
            # ("active_power", "kW", "active_power"),
            ("outdoor_temp", "°C", "outdoor_temp"),
            # ("frequency", "Hz", "frequency"),
            ("nacelle_pos", "°", "wind_direction"),  # renamed here
        ]

        # --------------------------------------------------
        # CREATE EXCEL
        # --------------------------------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = f"SCADA {report_label}"

        bold = Font(bold=True)

        headers = [
            "Latitude",
            "Longitude",
            "LocNo",
            "Machine",
            "Parameter",
            "Units",
            "Datetime"
        ]

        for i, h in enumerate(headers, start=1):
            ws.cell(row=i, column=1, value=h).font = bold

        col = 2

        # Header rows
        for locno in machines:
            machine_info = machine_data.get(locno, {})

            for field_name, unit, display_name in parameters:
                ws.cell(row=1, column=col, value=machine_info.get("latitude"))
                ws.cell(row=2, column=col, value=machine_info.get("longitude"))
                ws.cell(row=3, column=col, value=locno)
                ws.cell(row=4, column=col, value=locno)
                ws.cell(row=5, column=col, value=display_name)  # renamed only here
                ws.cell(row=6, column=col, value=unit)
                col += 1

        # --------------------------------------------------
        # WRITE DATA (NO CHANGE IN FIELD FETCHING)
        # --------------------------------------------------
        excel_row = 7

        for ts in tqdm(timestamps, desc="Writing Excel", unit="timestamps"):

            ws.cell(row=excel_row, column=1,
                    value=ts.strftime("%Y-%m-%d %H:%M"))

            col = 2

            for locno in machines:
                record = data_map[ts].get(locno)

                for field_name, _, _ in parameters:
                    value = getattr(record, field_name, None) if record else None
                    ws.cell(row=excel_row, column=col, value=value)
                    col += 1

            excel_row += 1

        # --------------------------------------------------
        # SAVE FILE
        # --------------------------------------------------
        reports_folder = os.path.join(settings.BASE_DIR, "reports")
        os.makedirs(reports_folder, exist_ok=True)

        file_name = f"scada_report_{report_label}.xlsx"
        path = os.path.join(reports_folder, file_name)

        wb.save(path)

        self.stdout.write(self.style.SUCCESS(f"\n✅ Report Generated Successfully"))
        self.stdout.write(self.style.SUCCESS(f"📁 Saved at: {path}\n"))